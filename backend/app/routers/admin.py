from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional
from datetime import datetime
import logging

from app.models.database import get_db
from app.models.entities import Admin, Interview, InterviewTask, AntiCheatEvent, TaskBank
from app.models.schemas import InterviewStatus, TaskType, Difficulty, InterviewDirection, ProgrammingLanguage
from fastapi import UploadFile, File, Form
from pydantic import BaseModel
from typing import List, Optional as Opt
import json
import csv
from io import StringIO, BytesIO
try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
from app.services.auth import (
    authenticate_admin, create_access_token, decode_token,
    get_admin_by_id, create_admin
)

router = APIRouter(prefix="/api/admin", tags=["admin"])
security = HTTPBearer()


async def get_current_admin(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
) -> Admin:
    """Verify JWT and return current admin"""
    token = credentials.credentials
    payload = decode_token(token)
    
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    admin_id = payload.get("sub")
    admin = await get_admin_by_id(db, admin_id)
    
    if not admin or not admin.is_active:
        raise HTTPException(status_code=401, detail="Admin not found or inactive")
    
    return admin


# Auth endpoints
@router.post("/login")
async def login(
    username: str,
    password: str,
    db: AsyncSession = Depends(get_db)
):
    """Admin login"""
    admin = await authenticate_admin(db, username, password)
    
    if not admin:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Update last login
    admin.last_login = datetime.utcnow()
    await db.commit()
    
    token = create_access_token({"sub": admin.id, "username": admin.username})
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "admin": {
            "id": admin.id,
            "username": admin.username,
            "email": admin.email,
            "is_superadmin": admin.is_superadmin
        }
    }


@router.post("/register")
async def register_admin(
    username: str,
    email: str,
    password: str,
    admin_key: str,  # Secret key to create new admins
    db: AsyncSession = Depends(get_db)
):
    """Register new admin (requires admin key)"""
    # Simple admin key check - in production use env var
    if admin_key != "super-secret-admin-key":
        raise HTTPException(status_code=403, detail="Invalid admin key")
    
    # Check if username exists
    existing = await db.execute(select(Admin).where(Admin.username == username))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Username already exists")
    
    admin = await create_admin(db, username, email, password)
    
    return {"message": "Admin created", "admin_id": admin.id}


# Dashboard endpoints
@router.get("/dashboard")
async def get_dashboard(
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Get admin dashboard statistics"""
    # Total interviews
    total_result = await db.execute(select(func.count(Interview.id)))
    total_interviews = total_result.scalar()
    
    # Completed interviews
    completed_result = await db.execute(
        select(func.count(Interview.id))
        .where(Interview.status == InterviewStatus.COMPLETED)
    )
    completed_interviews = completed_result.scalar()
    
    # Average scores
    avg_result = await db.execute(
        select(
            func.avg(Interview.overall_score),
            func.avg(Interview.technical_score),
            func.avg(Interview.softskills_score)
        ).where(Interview.status == InterviewStatus.COMPLETED)
    )
    avg_scores = avg_result.one()
    
    # Recent interviews
    recent_result = await db.execute(
        select(Interview)
        .order_by(Interview.created_at.desc())
        .limit(10)
    )
    recent_interviews = recent_result.scalars().all()
    
    # Flagged interviews (anti-cheat)
    flagged_result = await db.execute(
        select(func.count(Interview.id))
        .where(Interview.anti_cheat_flags > 2)
    )
    flagged_count = flagged_result.scalar()
    
    return {
        "total_interviews": total_interviews,
        "completed_interviews": completed_interviews,
        "in_progress": total_interviews - completed_interviews,
        "average_scores": {
            "overall": round(avg_scores[0] or 0, 1),
            "technical": round(avg_scores[1] or 0, 1),
            "softskills": round(avg_scores[2] or 0, 1)
        },
        "flagged_interviews": flagged_count,
        "recent_interviews": [
            {
                "id": i.id,
                "candidate_name": i.candidate_name,
                "direction": i.direction,  # direction теперь строка
                "status": i.status.value,
                "overall_score": i.overall_score,
                "created_at": i.created_at.isoformat()
            }
            for i in recent_interviews
        ]
    }


@router.get("/interviews")
async def list_interviews(
    status: Optional[str] = None,
    direction: Optional[str] = None,
    flagged_only: bool = False,
    page: int = 1,
    limit: int = 20,
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """List all interviews with filters"""
    query = select(Interview)
    
    if status:
        query = query.where(Interview.status == InterviewStatus(status))
    if direction:
        query = query.where(Interview.direction == direction)
    if flagged_only:
        query = query.where(Interview.anti_cheat_flags > 2)
    
    query = query.order_by(Interview.created_at.desc())
    query = query.offset((page - 1) * limit).limit(limit)
    
    result = await db.execute(query)
    interviews = result.scalars().all()
    
    # Get total count
    count_query = select(func.count(Interview.id))
    if status:
        count_query = count_query.where(Interview.status == InterviewStatus(status))
    if direction:
        count_query = count_query.where(Interview.direction == direction)
    if flagged_only:
        count_query = count_query.where(Interview.anti_cheat_flags > 2)
    
    count_result = await db.execute(count_query)
    total = count_result.scalar()
    
    return {
        "interviews": [
            {
                "id": i.id,
                "candidate_name": i.candidate_name,
                "candidate_email": i.candidate_email,
                "direction": i.direction,  # direction теперь строка
                "difficulty": i.difficulty.value,
                "status": i.status.value,
                "overall_score": i.overall_score,
                "technical_score": i.technical_score,
                "softskills_score": i.softskills_score,
                "tasks_completed": i.tasks_completed,
                "total_tasks": i.total_tasks,
                "hints_used": i.hints_used,
                "anti_cheat_flags": i.anti_cheat_flags,
                "recommendation": i.recommendation,
                "created_at": i.created_at.isoformat(),
                "finished_at": i.finished_at.isoformat() if i.finished_at else None
            }
            for i in interviews
        ],
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }


@router.get("/interviews/{interview_id}")
async def get_interview_details(
    interview_id: str,
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Get detailed interview report"""
    result = await db.execute(select(Interview).where(Interview.id == interview_id))
    interview = result.scalar_one_or_none()
    
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    
    # Get tasks
    tasks_result = await db.execute(
        select(InterviewTask)
        .where(InterviewTask.interview_id == interview.id)
        .order_by(InterviewTask.task_number)
    )
    tasks = tasks_result.scalars().all()
    
    # Get anti-cheat events
    events_result = await db.execute(
        select(AntiCheatEvent)
        .where(AntiCheatEvent.interview_id == interview.id)
        .order_by(AntiCheatEvent.created_at)
    )
    events = events_result.scalars().all()
    
    return {
        "interview": {
            "id": interview.id,
            "candidate_name": interview.candidate_name,
            "candidate_email": interview.candidate_email,
            "direction": interview.direction,  # direction теперь строка
            "language": interview.language.value,
            "difficulty": interview.difficulty.value,
            "status": interview.status.value,
            "overall_score": interview.overall_score,
            "technical_score": interview.technical_score,
            "softskills_score": interview.softskills_score,
            "strengths": interview.strengths,
            "areas_for_improvement": interview.areas_for_improvement,
            "recommendation": interview.recommendation,
            "softskills_assessment": interview.softskills_assessment,
            "hints_used": interview.hints_used,
            "anti_cheat_flags": interview.anti_cheat_flags,
            "started_at": interview.started_at.isoformat() if interview.started_at else None,
            "finished_at": interview.finished_at.isoformat() if interview.finished_at else None
        },
        "tasks": [
            {
                "id": t.id,
                "task_number": t.task_number,
                "title": t.title,
                "description": t.description,
                "task_type": t.task_type.value,
                "submitted_code": t.submitted_code,
                "score": t.score,
                "feedback": t.feedback,
                "code_quality": t.code_quality,
                "efficiency": t.efficiency,
                "correctness": t.correctness,
                "hints_used": t.hints_used,
                "execution_result": t.execution_result
            }
            for t in tasks
        ],
        "anti_cheat_events": [
            {
                "type": e.event_type.value,
                "severity": e.severity,
                "details": e.details,
                "timestamp": e.created_at.isoformat()
            }
            for e in events
        ]
    }


@router.delete("/interviews/{interview_id}")
async def delete_interview(
    interview_id: str,
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Delete an interview (superadmin only)"""
    if not admin.is_superadmin:
        raise HTTPException(status_code=403, detail="Superadmin required")
    
    result = await db.execute(select(Interview).where(Interview.id == interview_id))
    interview = result.scalar_one_or_none()
    
    if not interview:
        raise HTTPException(status_code=404, detail="Interview not found")
    
    await db.delete(interview)
    await db.commit()
    
    return {"message": "Interview deleted"}


# Task Bank endpoints
@router.post("/task-bank/upload")
async def upload_task_bank(
    file: UploadFile = File(...),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Upload tasks from CSV or JSON file"""
    content = await file.read()
    file_content = content.decode('utf-8')
    
    tasks = []
    if file.filename.endswith('.json'):
        tasks = json.loads(file_content)
    elif file.filename.endswith('.csv'):
        # Parse CSV
        reader = csv.DictReader(StringIO(file_content))
        for row in reader:
            task = {
                "title": row.get("title", ""),
                "description": row.get("description", ""),
                "task_type": row.get("task_type", "algorithm"),
                "difficulty": row.get("difficulty", "medium"),
                "direction": row.get("direction", "backend"),
                "examples": json.loads(row.get("examples", "[]")) if row.get("examples") else [],
                "constraints": json.loads(row.get("constraints", "[]")) if row.get("constraints") else [],
                "test_cases": json.loads(row.get("test_cases", "[]")) if row.get("test_cases") else [],
                "starter_code": json.loads(row.get("starter_code", "{}")) if row.get("starter_code") else {},
                "topic": row.get("topic", ""),
                "tags": json.loads(row.get("tags", "[]")) if row.get("tags") else [],
            }
            tasks.append(task)
    elif file.filename.endswith(('.xlsx', '.xls')):
        if not HAS_XLSX:
            raise HTTPException(status_code=400, detail="XLSX support requires openpyxl. Install with: pip install openpyxl")
        # Parse XLSX
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            task = {
                "title": str(row_dict.get("title", "")),
                "description": str(row_dict.get("description", "")),
                "task_type": str(row_dict.get("task_type", "algorithm")),
                "difficulty": str(row_dict.get("difficulty", "medium")),
                "direction": str(row_dict.get("direction", "backend")),
                "examples": json.loads(row_dict.get("examples", "[]")) if row_dict.get("examples") else [],
                "constraints": json.loads(row_dict.get("constraints", "[]")) if row_dict.get("constraints") else [],
                "test_cases": json.loads(row_dict.get("test_cases", "[]")) if row_dict.get("test_cases") else [],
                "starter_code": json.loads(row_dict.get("starter_code", "{}")) if row_dict.get("starter_code") else {},
                "topic": str(row_dict.get("topic", "")),
                "tags": json.loads(row_dict.get("tags", "[]")) if row_dict.get("tags") else [],
            }
            tasks.append(task)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV, JSON, or XLSX")
    
    # Generate embeddings and save tasks
    from app.services.scibox_client import scibox_client
    saved_count = 0
    
    for task_data in tasks:
        # Create task in bank
        task = TaskBank(
            title=task_data["title"],
            description=task_data["description"],
            task_type=TaskType(task_data["task_type"]),
            difficulty=Difficulty(task_data["difficulty"]),
            direction=InterviewDirection(task_data["direction"]),
            examples=task_data.get("examples"),
            constraints=task_data.get("constraints"),
            test_cases=task_data.get("test_cases"),
            starter_code=task_data.get("starter_code"),
            topic=task_data.get("topic"),
            tags=task_data.get("tags"),
            created_by=admin.id
        )
        
        # Generate embedding for task description + topic
        embedding_text = f"{task_data['title']} {task_data['description']} {task_data.get('topic', '')}"
        try:
            embeddings = await scibox_client.get_embeddings([embedding_text])
            if embeddings:
                task.embedding = embeddings[0]
        except Exception as e:
            print(f"Failed to generate embedding for task {task_data['title']}: {e}")
        
        db.add(task)
        saved_count += 1
    
    await db.commit()
    
    return {
        "message": f"Successfully uploaded {saved_count} tasks",
        "count": saved_count
    }


@router.get("/task-bank")
async def list_task_bank(
    difficulty: Optional[str] = None,
    direction: Optional[str] = None,
    task_type: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """List tasks from bank"""
    query = select(TaskBank)
    
    if difficulty:
        query = query.where(TaskBank.difficulty == Difficulty(difficulty))
    if direction:
        query = query.where(TaskBank.direction == InterviewDirection(direction))
    if task_type:
        query = query.where(TaskBank.task_type == TaskType(task_type))
    
    query = query.order_by(TaskBank.created_at.desc())
    query = query.offset((page - 1) * limit).limit(limit)
    
    result = await db.execute(query)
    tasks = result.scalars().all()
    
    # Get total count
    count_query = select(func.count(TaskBank.id))
    if difficulty:
        count_query = count_query.where(TaskBank.difficulty == Difficulty(difficulty))
    if direction:
        count_query = count_query.where(TaskBank.direction == InterviewDirection(direction))
    if task_type:
        count_query = count_query.where(TaskBank.task_type == TaskType(task_type))
    
    count_result = await db.execute(count_query)
    total = count_result.scalar()
    
    return {
        "tasks": [
            {
                "id": t.id,
                "title": t.title,
                "description": t.description,
                "task_type": t.task_type.value,
                "difficulty": t.difficulty.value,
                "direction": t.direction.value,
                "topic": t.topic,
                "tags": t.tags,
                "times_used": t.times_used,
                "average_score": t.average_score,
                "created_at": t.created_at.isoformat()
            }
            for t in tasks
        ],
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }


@router.post("/task-bank/single")
async def add_single_task(
    description: str = Form(...),
    difficulty: str = Form(...),
    task_type: str = Form("algorithm"),
    title: Opt[str] = Form(None),
    direction: Opt[str] = Form("backend"),
    expected_solution: Opt[str] = Form(None),
    examples: Opt[str] = Form(None),
    constraints: Opt[str] = Form(None),
    test_cases: Opt[str] = Form(None),
    starter_code: Opt[str] = Form(None),
    topic: Opt[str] = Form(None),
    tags: Opt[str] = Form(None),
    language: Opt[str] = Form(None),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Add a single task to bank (3 required fields: description, difficulty, task_type)"""
    try:
        # Validate and parse enums
        try:
            task_type_enum = TaskType(task_type)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid task_type: {task_type}. Must be one of: algorithm, system_design, code_review, debugging, practical"
            )
        
        try:
            difficulty_enum = Difficulty(difficulty)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid difficulty: {difficulty}. Must be one of: easy, medium, hard"
            )
        
        try:
            direction_enum = InterviewDirection(direction) if direction else InterviewDirection.BACKEND
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid direction: {direction}. Must be one of: frontend, backend, fullstack, data_science, devops"
            )
        
        language_enum = None
        if language:
            try:
                language_enum = ProgrammingLanguage(language)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid language: {language}. Must be one of: python, javascript, cpp"
                )
        
        # Parse JSON fields with error handling
        examples_parsed = None
        if examples:
            try:
                examples_parsed = json.loads(examples)
            except json.JSONDecodeError as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid JSON in examples field: {str(e)}"
                )
        
        constraints_parsed = None
        if constraints:
            try:
                constraints_parsed = json.loads(constraints)
            except json.JSONDecodeError as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid JSON in constraints field: {str(e)}"
                )
        
        test_cases_parsed = None
        if test_cases:
            try:
                test_cases_parsed = json.loads(test_cases)
            except json.JSONDecodeError as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid JSON in test_cases field: {str(e)}"
                )
        
        starter_code_parsed = None
        if starter_code:
            try:
                starter_code_parsed = json.loads(starter_code)
            except json.JSONDecodeError as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid JSON in starter_code field: {str(e)}"
                )
        
        tags_parsed = None
        if tags:
            try:
                tags_parsed = json.loads(tags)
            except json.JSONDecodeError as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid JSON in tags field: {str(e)}"
                )
        
        # Create task
        task = TaskBank(
            title=title or f"Задача {datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
            description=description,
            task_type=task_type_enum,
            difficulty=difficulty_enum,
            direction=direction_enum,
            expected_solution=expected_solution,
            examples=examples_parsed,
            constraints=constraints_parsed,
            test_cases=test_cases_parsed,
            starter_code=starter_code_parsed,
            topic=topic,
            tags=tags_parsed,
            language=language_enum,
            created_by=admin.id
        )
        
        # Generate embedding
        from app.services.scibox_client import scibox_client
        embedding_text = f"{task.title} {task.description} {task.topic or ''}"
        try:
            embeddings = await scibox_client.get_embeddings([embedding_text])
            if embeddings:
                task.embedding = embeddings[0]
        except Exception as e:
            logging.warning(f"Failed to generate embedding: {e}")
            # Continue without embedding - it's not critical
        
        try:
            db.add(task)
            await db.commit()
            await db.refresh(task)
        except Exception as db_error:
            await db.rollback()
            logging.error(f"Database error adding task: {db_error}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Database error: {str(db_error)}"
            )
        
        return {
            "message": "Task added successfully",
            "task_id": task.id
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error adding task: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Internal server error: {str(e)}"
        )


@router.get("/task-bank/export")
async def export_task_bank(
    task_ids: Opt[str] = None,  # Comma-separated IDs, or "all"
    format: str = "json",  # json, csv, xlsx
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Export tasks from bank"""
    if task_ids == "all" or not task_ids:
        query = select(TaskBank)
    else:
        ids_list = task_ids.split(",")
        query = select(TaskBank).where(TaskBank.id.in_(ids_list))
    
    result = await db.execute(query)
    tasks = result.scalars().all()
    
    if format == "json":
        tasks_data = [
            {
                "id": t.id,
                "title": t.title,
                "description": t.description,
                "task_type": t.task_type.value,
                "difficulty": t.difficulty.value,
                "direction": t.direction.value,
                "examples": t.examples,
                "constraints": t.constraints,
                "test_cases": t.test_cases,
                "starter_code": t.starter_code,
                "expected_solution": t.expected_solution,
                "topic": t.topic,
                "tags": t.tags,
            }
            for t in tasks
        ]
        from fastapi.responses import JSONResponse
        return JSONResponse(content=tasks_data)
    
    elif format == "csv":
        from fastapi.responses import Response
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "title", "description", "task_type", "difficulty", "direction",
            "examples", "constraints", "test_cases", "starter_code", "topic", "tags"
        ])
        writer.writeheader()
        for t in tasks:
            writer.writerow({
                "title": t.title,
                "description": t.description,
                "task_type": t.task_type.value,
                "difficulty": t.difficulty.value,
                "direction": t.direction.value,
                "examples": json.dumps(t.examples) if t.examples else "",
                "constraints": json.dumps(t.constraints) if t.constraints else "",
                "test_cases": json.dumps(t.test_cases) if t.test_cases else "",
                "starter_code": json.dumps(t.starter_code) if t.starter_code else "",
                "topic": t.topic or "",
                "tags": json.dumps(t.tags) if t.tags else "",
            })
        return Response(content=output.getvalue(), media_type="text/csv", headers={
            "Content-Disposition": "attachment; filename=tasks_export.csv"
        })
    
    elif format == "xlsx":
        if not HAS_XLSX:
            raise HTTPException(status_code=400, detail="XLSX export requires openpyxl")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["title", "description", "task_type", "difficulty", "direction",
                   "examples", "constraints", "test_cases", "starter_code", "topic", "tags"])
        for t in tasks:
            ws.append([
                t.title,
                t.description,
                t.task_type.value,
                t.difficulty.value,
                t.direction.value,
                json.dumps(t.examples) if t.examples else "",
                json.dumps(t.constraints) if t.constraints else "",
                json.dumps(t.test_cases) if t.test_cases else "",
                json.dumps(t.starter_code) if t.starter_code else "",
                t.topic or "",
                json.dumps(t.tags) if t.tags else "",
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        from fastapi.responses import Response
        return Response(content=output.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       headers={"Content-Disposition": "attachment; filename=tasks_export.xlsx"})
    
    else:
        raise HTTPException(status_code=400, detail="Unsupported format. Use json, csv, or xlsx")


@router.delete("/task-bank/{task_id}")
async def delete_task_from_bank(
    task_id: str,
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Delete a task from bank"""
    result = await db.execute(select(TaskBank).where(TaskBank.id == task_id))
    task = result.scalar_one_or_none()
    
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    await db.delete(task)
    await db.commit()
    
    return {"message": "Task deleted"}


@router.delete("/task-bank")
async def delete_multiple_tasks(
    task_ids: str,  # Comma-separated IDs
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Delete multiple tasks from bank"""
    ids_list = task_ids.split(",")
    result = await db.execute(select(TaskBank).where(TaskBank.id.in_(ids_list)))
    tasks = result.scalars().all()
    
    for task in tasks:
        await db.delete(task)
    
    await db.commit()
    
    return {"message": f"Deleted {len(tasks)} tasks"}
